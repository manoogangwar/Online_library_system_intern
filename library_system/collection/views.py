import openpyxl
from django.http import HttpResponse
from django.shortcuts import render,redirect
from django.views.generic import CreateView,ListView,View
import openpyxl.workbook
from .models import author,book,borrowRecord
from .forms import authorForm,bookForm,borrowRecordForm
from django.urls import reverse_lazy
from django.contrib import messages

# Create your views here.

def exportView(request):
    return render(request, 'download_button.html')


class addingAuthorView(CreateView):
    model = author
    form_class = authorForm
    template_name = 'author_form.html'
    success_url = reverse_lazy('author_list')
    
    def from_valid(self,form):
        messages.success(self.request, 'author added successfully')
        return super().form_valid(form)
    
    
class addBookView(CreateView):
    model = book
    form_class = bookForm
    template_name = 'book_form.html'
    success_url = reverse_lazy('book_list')
    
    
    def form_valid(self,form):
        messages.success(self.request,'Book added successfully')
        return super().form_valid(form)
    
    
class addborrowRecordView(CreateView):
    model = borrowRecord
    form_class = borrowRecordForm
    template_name = 'borrowRecord_form.html'
    success_url = reverse_lazy('borrow_list')
    
    def form_valid(self,form):
        messages.success(self.request, 'barrow record added successfully')
        return super().form_valid(form)
    
    
    
    
class authorListView(ListView):
    model = author
    template_name = 'author_list.html'
    context_object_name = 'authors'
    paginate_by = 15


class bookListView(ListView):
    model = book
    template_name = 'book_list.html'
    context_object_name = 'books'
    paginate_by = 15


class borrowRecordListView(ListView):
    model = borrowRecord
    template_name = 'borrow_list.html'
    context_object_name = 'records'
    paginate_by = 15



class exportExelView(View):
    def get(self,request,*args,**kwargs):
        workbook = openpyxl.Workbook()
        
        author_sheet = workbook.active
        author_sheet.title = 'authors'
        author_sheet.append(['id','name','email','bio'])
        for auth in author.objects.all():
             author_sheet.append([auth.id, auth.name, auth.email, auth.bio])
            
            
        book_sheet=workbook.create_sheet(title='books')
        book_sheet.append(['id','book.title','genre','published_date','author'])
        for b in book.objects.all():
            book_sheet.append([b.id, b.title, b.genre, b.published_date, b.author.name])
            
            
        record_sheet = workbook.create_sheet(title='borrowRecord' )
        record_sheet.append(['id','book','borrower','borrow_date','return_date'])
        for record in borrowRecord.objects.all():
            record_sheet.append([record.id,record.borrow_name,record.book.title,record.borrow_date,record.return_date])
            
            
            
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=library_data.xlsx'
        workbook.save(response)
        return response